import openpyxl


class Member:

    def __init__(self, first: str, last: str, phone: str, email: str, status: str, active: bool):
        """
        Holds metadata and real data for members of the fraternity.
        Stored in the hours field is the running total of cleanup hours
        the member currently has. Default to -1 for work with exception
        handling
        :param first: first name of member
        :param last: last name of member
        :param phone: phone number of member
        :param email: email address of member
        :param status: whether the member is an Associate Member,
                       a Brother, or part of the Sayonara Squad (SS)
        :param active: whether the member is and active brother
        """
        self.first = first
        self.last = last
        self.phone = phone
        self.email = email
        self.status = status
        self.active = active
        self.hours = -1

    def __str__(self):
        return """Name: """ + self.first + ' ' + self.last\
               + """\nPhone: """ + str(self.phone)\
               + """\nEmail: """ + str(self.email)\
               + """\nStatus: """ + str(self.status)\
               + """\nActive: """ + str(self.active)\
               + """\nHours: """ + str(self.hours) + '\n'

    def member_status(self) -> int:
        """
        Simple mechanism to providing a comparable interface between members
        based on status. Lower number implies greater importance. To be used
        as a key in a lambda expression to perform sorting on lists of Members
        :rtype: int
        :return: integer equivalent of the status of the brother.
        TODO: fill out AM
        """
        if self.status == 'SS':
            return 0
        elif self.status == 'NIB':
            return 2
        else:
            return 1


def generate_members() -> list:
    """
    Generates a list of Member objects from Cleanup Sheet.xlsx. Each member's
    data is extracted and stored in a Member. Running total of hours is accessed by
    calling the generate_hours method
    :rtype: list
    :return: list of Member objects
    """
    excel_document = openpyxl.load_workbook('Cleanup Sheet.xlsx')
    member_sheet = excel_document['Members']
    hours_dict = generate_hours()
    members = []
    for i in range(2, member_sheet.max_row + 1):
        member_row = member_sheet[i]
        name_first = member_row[0].value
        name_last = member_row[1].value
        phone = member_row[2].value
        email = member_row[3].value
        status = member_row[4].value
        active = member_row[5].value
        member = Member(name_first, name_last, phone, email, status, active)
        member.hours = hours_dict.get(name_first.strip() + name_last.strip(), -1)
        members.append(member)
        # print(member)
    return members


def generate_hours() -> dict:
    """
    Retrieves the running total of hours for all members from the Cleanup Sheet.xlsx
    The values are stored in a dictionary where the key is the member's last name
    :rtype: dict
    :return: dict of hours mapping value 'hours' to key 'last name'
    """
    excel_document = openpyxl.load_workbook('Cleanup Sheet.xlsx')
    member_sheet = excel_document['Running Total']
    hours_dict = {}
    for i in range(2, member_sheet.max_row + 1):
        member_row = member_sheet[i]
        hours_dict[member_row[0].value.strip() + member_row[1].value.strip()] = member_row[2].value
    return hours_dict


if __name__ == '__main__':
    # To see output, uncomment print statements in generate_members()
    generate_members()
